import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
import json
import time

#test buy, sell close:

#from tbuy import buy
#from tclose import close
from tsell import sell

#actul buy, sell, close(comment out to run test):

from buy import buy
from close import close
#from sell import sell()

TIME = 53
#Indicatori:
SMA = 7
LMA = 24
PVTW = 10
HGW = 10
#Setari Pozitii:
STOPLOSSP = 0.9
TAKEPROFITP= 1
TSP = STOPLOSSP * 0.001 #teiling stoploss, the formula i use is 
TTP = TAKEPROFITP * 0.002
SP = 0.003
TP = 0.0015

SP_P = 0.002
ATRP = 14

json_paths = [
    r'C:\Users\David\Desktop\Fundamental_analysis\json_prices\price_1.json',
    r'C:\Users\David\Desktop\Fundamental_analysis\json_prices\price_2.json',
    r'C:\Users\David\Desktop\Fundamental_analysis\json_prices\price_3.json',
    r'C:\Users\David\Desktop\Fundamental_analysis\json_prices\price_4.json',
    r'C:\Users\David\Desktop\Fundamental_analysis\json_prices\price_5.json',
    r'C:\Users\David\Desktop\Fundamental_analysis\json_prices\price_6.json'
]
def profit_save(number):
    with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\proft_test.txt', 'a') as file:
        file.write(f"{number} + ")

def pivot_points(series, window):
    max_val = np.max(series)
    min_val = np.min(series)
    last_val = series[-1]
    
    if last_val == max_val:
        return 1
    elif last_val == min_val:
        return -1
    else:
        return 0


def read_values_from_j(JSON_FILE_PATH):
    with open(JSON_FILE_PATH, "r") as json_file:
        try:
            data = json.load(json_file)
        except json.decoder.JSONDecodeError as e:
            print(f"Error decoding JSON at line {e.lineno}, column {e.colno}: {e.msg}")
            raise
    return data

def update_excel_and_save_to_j(data):
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
        sheet.delete_rows(1, sheet.max_row)
        for item in data:
            next_row = sheet.max_row + 1
            sheet.cell(row=next_row, column=1, value=item["timestamp"])
            sheet.cell(row=next_row, column=2, value=item["price"])
        workbook.save(EXCEL_FILE_PATH)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

def tp_json():
    try:
        with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\TakeProfit.json', 'r') as json_file:
            data = json.load(json_file)
            takeprofit = data.get("take_profit")
            return takeprofit
    except FileNotFoundError:
        print("TakeProfit.json not found. Returning None.")
        return None

def tp_jsonSELL():
    try:
        with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\TakeProfitSELL.json', 'r') as json_file:
            data = json.load(json_file)
            takeProfitSELL = data.get("take_profit")
            return takeProfitSELL
    except FileNotFoundError:
        print("TakeProfitSELL.json not found. Returning None.")
        return None
    
def sp_json():
    try:
        with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\StopLoss.json', 'r') as json_file:
            data = json.load(json_file)
            stoploss = data.get("stop_loss")
            return stoploss
    except FileNotFoundError:
        print("StopLoss.json not found. Returning None.")
        return None

def sp_jsonSELL():
    try:
        with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\StopLossSELL.json', 'r') as json_file:
            data = json.load(json_file)
            stopLossSELL = data.get("stop_loss")
            return stopLossSELL
    except FileNotFoundError:
        print("StopLossSELL.json not found. Returning None.")
        return None
'''
# Exemplu:
j_data = read_values_from_j(j_FILE_PATH)
update_excel_and_save_to_j(j_data)
'''
print('STARTING PROFIT PILOT...')
time.sleep(2)

pos = input("Buy position opened(y/n/s): ")
if pos.lower() == 'y':
    pos = 1
if pos.lower() == 'n':
    pos = 0


timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
print(f'START { timestamp }')
lstpos = 0

def main(ticker, j_dta, pos):
    pos = pos
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


    df['SMA'] = df['Close'].rolling(SMA).mean()
    df['LMA'] = df['Close'].rolling(LMA).mean()
    df['SMAvLMA'] = np.where(df['SMA'] > df['LMA'], 1, 0)
    df['ClosevMA20'] = np.where(df['Close'] > df['SMA'], 1, 0)
    df['Conclusion'] = np.where((df['SMAvLMA'] >= 1) & (df['ClosevMA20'] >= 1), 1, 0)
    df['MACD'] = df['SMA'] - df['LMA']
    df['SL'] = df['MACD'].rolling(window=9, min_periods=1).mean()
    df['High'] = df['Close'].shift(1).cummax()  
    df['Low'] = df['Close'].shift(1).cummin()
    '''df['TR'] = np.maximum.reduce([
        df['High'] - df['Low'],
        np.abs(df['High'] - df['Close'].shift(1)),
        np.abs(df['Low'] - df['Close'].shift(1))
    ])
    df['ATR'] = df['TR'].rolling(window=ATRP).mean()'''

    #df["stopLoss"] = df["Close"] - SP * df["ATR"]
    #df["takeProfit"] = df["Close"] + TP * df["ATR"]

    #df["high_pivot"] = df["High"].rolling(PVTW).apply(lambda x: pivot_points(x, PVTW), raw=True)
    #df["low_pivot"] = df["Low"].rolling(PVTW).apply(lambda x: pivot_points(x, PVTW), raw=True)

    #df["pivot"] = df["high_pivot"] + df["low_pivot"]
    #df["pivot"] = df["pivot"].replace(0, pd.NA).dropna()

    #df["high_of_trend"] = df['Close'].tail(10).max()
    #df["low_of_trend"] = df['Close'].tail(10).min()

    if pos == 0:
        openPrice = 25000
    #CONDITION:
    BuyCondition = False  
    SellCondition = False
    CloseBuy = False
    CloseSell = False
    #CONDITION  - BUY
    #and (df['MACD'].iloc[-1] > df['SL'].iloc[-1]).any()
    if(df['SMA'].iloc[-1] > df['LMA'].iloc[-1]).any()  and pos == 0:
        BuyCondition = True
    
    #CONDITION - SELL
    #and (df['MACD'].iloc[-1] < df['SL'].iloc[-1]).any()
    if (df['SMA'].iloc[-1] < df['LMA'].iloc[-1]).any()  and pos == 0:
        SellCondition = True
    
    #CONDITION - CLOSE BUY
    #and (df['MACD'].iloc[-1] < df['SL'].iloc[-1]).any()
    if (df['SMA'].iloc[-1] < df['LMA'].iloc[-1]).any()  and pos == 1:
        CloseBuy = True
    
    #CONDITION - CLOSE SELL
    #and (df['MACD'].iloc[-1] < df['SL'].iloc[-1]).any()
    if (df['SMA'].iloc[-1] < df['LMA'].iloc[-1]).any()  and pos == 1:
        CloseSell = True

#    BUY
    if BuyCondition is not None and BuyCondition == True:
        stopLoss = df['stopLoss'].iloc[-1]
        takeProfit = df['takeProfit'].iloc[-1]
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f' BUY - time: {timestamp}, Close: {price}, MAL: {lma}, MS: {sma}')
        buy()
        pos = 1
        buyOpenPrice = df['Close'].iloc[-1] 
        BuyCondition = False

#    SELL
    if SellCondition is not None and SellCondition == True :
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f' SELL - time: {timestamp}, Close: {price}, MAL: {lma}, MS: {sma}, stopLoss: {stopLossSELL}, TP: {takeProfitSELL}')
        #sell()
        pos = 2
        sellOpenPrice = df['Close'].iloc[-1]
        SellCondition = False

#   CONDITION CLOSE BUY   
    if CloseBuy is not None and CloseBuy == True:
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        close()
        ClosePrice = price
        BuyProfit = ((ClosePrice - buyOpenPrice)/ buyOpenPrice)* 100
        print(f'SMA CLOSE ({BuyProfit}) - time: {timestamp}, Close: {price}, MAL: {lma}, MS: {sma}') 
        pos = 0
        lstpos = 1
        profit_save(BuyProfit)
        CloseBuy = False
    
#   CONDITION CLOSE SELL  
    if CloseSell is not None and CloseSell == True :
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ClosePrice = price
        SellProfit = "NaN" #((sellOpenPrice - ClosePrice)/ sellOpenPrice)* 100
        print(f'SMA CLOSE ({SellProfit}) - time: {timestamp}, Close: {price}, MAL: {lma}, MS: {sma}') 
        close()
        pos = 0
        lstpos = 2
        profit_save(SellProfit)
        CloseSell = False
#   OPEN BUY
   # Cprice = df['Close'].iloc[-1]
    if df['SMA'].iloc[-1] > df['LMA'].iloc[-1] and pos == 1:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]  
       # stopLoss = stopLoss + ( stopLoss*TSP)
        #takeProfit = takeProfit + ( takeProfit * TTP)
        pos = 1
        print(f'OPEN BUY - time: {timestamp}, Close: {price}, MAL: {lma}, MS: {sma}, stopLoss: {stopLoss}, ')

#   OPEN SELL
    Cprice = df['Close'].iloc[-1]
    if df['SMA'].iloc[-1] < df['LMA'].iloc[-1] and pos == 2:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]  
        #stopLossSELL = stopLossSELL - ( stopLoss*TSP)
        #takeProfitSELL = takeProfitSELL - ( takeProfit * TTP)
        stopLoss = "NaN"
        takeProfit = "NaN"
        pos = 2
        print(f'OPEN SELL - time: {timestamp}, Close: {price}, MAL: {lma}, MS: {sma}, stopLoss: {stopLoss}, ')

#   NO ACTION
    if pos == 0:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]
        print(f'NO ACTION - time: {timestamp}, Close: {price}, MAL: {lma}, MS: {sma}')
        SELL_reentry_condition = 0
        BUY_reentry_condition = 0
        # Re-enter BUY position
        if df['SMA'].iloc[-1] > df['SMA'].iloc[-1] and df['Close'].iloc[-1] >= df["high_of_trend"].iloc[-1]:
            BUY_reentry_condition = 1
            SELL_reentry_condition = 0
        if pos == 0 and BUY_reentry_condition == 1:
            if df['Close'].iloc[-1] >= df['high_of_trend'].iloc[-1]:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                price = df['Close'].iloc[-1]
                lma = df['LMA'].iloc[-1]
                sma = df['SMA'].iloc[-1]
                pos = 1
                buyOpenPrice = df['Close'].iloc[-1]
                print(f'RE-ENTER BUY - time: {timestamp}, Close: {price}, MAL: {lma}, MS: {sma}, stopLoss: {stopLoss}')
                buy()
                BUY_reentry_condition = 0
    while pos == 1:
        main(ticker, j_data, pos)
        time.sleep(TIME)
