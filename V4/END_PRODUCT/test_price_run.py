import json
import pandas as pd
import time
import datetime
from datetime import datetime
import openpyxl

TIME = 4
EXCEL_INTAKE = r'C:\Users\David\Desktop\Pilot\END_PRODUCT\TEST_DATA.xlsx'
J_FILE_PATH = r'C:\Users\David\Desktop\Pilot\END_PRODUCT\history.json'
NUMBER_OF_ROWS = 500
MAX_HISTORY_ENTRIES = 80

def update_excel_and_save_to_j(data):
    try:
        workbook = openpyxl.load_workbook(EXCEL_INTAKE)
        sheet = workbook.active
        sheet.delete_rows(1, sheet.max_row)
        for item in data:
            next_row = sheet.max_row + 1
            sheet.cell(row=next_row, column=1, value=item["timestamp"])
            sheet.cell(row=next_row, column=2, value=item["price"])
        workbook.save(EXCEL_INTAKE)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

def read_values_from_j(JSON_FILE_PATH):
    with open(JSON_FILE_PATH, "r") as json_file:
        try:
            data = json.load(json_file)
        except json.decoder.JSONDecodeError as e:
            print(f"Error decoding JSON at line {e.lineno}, column {e.colno}: {e.msg}")
            raise
    return data

def save_price_to_data(price):
    json_file = r'C:\Users\David\Desktop\Pilot\END_PRODUCT\data_test.json'
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data = {"price": price, "timestamp": timestamp}
    
    try:
        with open(json_file, "r") as file:
            history_data = json.load(file)
    except (FileNotFoundError, json.decoder.JSONDecodeError):
        history_data = []

    history_data.append(data)

    # Keep only the last MAX_HISTORY_ENTRIES entries
    history_data = history_data[-MAX_HISTORY_ENTRIES:]

    with open(json_file, "w") as file:
        json.dump(history_data, file, indent=2)

j_data = read_values_from_j(J_FILE_PATH)
update_excel_and_save_to_j(j_data)
df  = df = pd.read_excel(EXCEL_INTAKE, names=['Date', 'Close'])
df1 = df.tail(NUMBER_OF_ROWS)

n = 0
while n < NUMBER_OF_ROWS:
    price = df1['Close'].iloc[n]
    n = n + 1
    save_price_to_data(price)
    print(f"running, price: { price }, {n}")
    time.sleep(TIME)