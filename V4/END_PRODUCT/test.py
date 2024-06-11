import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
import json
import time

#test buy, sell close:

from tbuy import buy
from tclose import close
from tsell import sell

#actul buy, sell, close(comment out to run test):

#from buy import buy
#from close import close
#from sell import sell()

TIME = 4
#Indicatori:
PVTW = 10
HGW = 10
SMA  = 5
LMA = 20
#Setari Pozitii:
STOPLOSSP = 0.9
TAKEPROFITP= 1
TSP = STOPLOSSP * 0.001 #teiling stoploss, the formula i use is 
TTP = TAKEPROFITP * 0.002
SP = 0.003
TP = 0.0015

SP_P = 0.002
ATRP = 14

EXCEL_FILE_PATH = r"C:\Users\David\Desktop\Pilot\END_PRODUCT\price.xlsx"
J_FILE_PATH = r"C:\Users\David\Desktop\Pilot\END_PRODUCT\data_test.json"

def profit_save(number):
    with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\proft_test.txt', 'a') as file:
        file.write(f"{number} + ")

def pivot_points(series, window):
    max_val = np.max(series)
    min_val = np.min(series)
    last_val = series[-1]
    
    if last_val == max_val:
        return 1
    elif last_val == min_val:
        return -1
    else:
        return 0
    
def stp(number):
    data = {"stop_loss": number}
    print(f"New StopLoss: { number }")
    with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\StopLoss.json', 'w') as json_file:
        json.dump(data, json_file)

def tke(number):
    data = {"take_profit": number}
    with open(r"C:\Users\David\Desktop\Pilot\END_PRODUCT\TakeProfit.json", 'w') as json_file:
        json.dump(data, json_file)

def stpSELL(number):
    data = {"stop_loss": number}
    print(f"New StopLoss: { number }")              #
    with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\StopLossSELL.json', 'w') as json_file:
        json.dump(data, json_file)

def tkeSELL(number):
    data = {"take_profit": number}
    with open(r"C:\Users\David\Desktop\Pilot\END_PRODUCT\TakeProfitSELL.json", 'w') as json_file:
        json.dump(data, json_file)

def find_optimal_lma_sma(df):
    optimal_lma = 0
    optimal_sma = 0
    max_sma_lma_price_sum = float('-inf')

    for lma in range(5, 26):
        for sma in range(5, lma - 4):
            df['SMA'] = df['Close'].rolling(sma).mean()
            df['LMA'] = df['Close'].rolling(lma).mean()
            df['SMA-LMAPrice'] = (df['Close'] / df['Close'].shift(1) - 1).where(df['SMA'] > df['LMA']) * 100

            sma_lma_price_sum = df['SMA-LMAPrice'].sum()

            if sma_lma_price_sum > max_sma_lma_price_sum:
                optimal_lma = lma
                optimal_sma = sma
                max_sma_lma_price_sum = sma_lma_price_sum

    return optimal_lma, optimal_sma, max_sma_lma_price_sum

def read_values_from_j(JSON_FILE_PATH):
    with open(JSON_FILE_PATH, "r") as json_file:
        try:
            data = json.load(json_file)
        except json.decoder.JSONDecodeError as e:
            print(f"Error decoding JSON at line {e.lineno}, column {e.colno}: {e.msg}")
            raise
    return data

def update_excel_and_save_to_j(data):
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
        sheet.delete_rows(1, sheet.max_row)
        for item in data:
            next_row = sheet.max_row + 1
            sheet.cell(row=next_row, column=1, value=item["timestamp"])
            sheet.cell(row=next_row, column=2, value=item["price"])
        workbook.save(EXCEL_FILE_PATH)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

def tp_json():
    try:
        with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\TakeProfit.json', 'r') as json_file:
            data = json.load(json_file)
            takeprofit = data.get("take_profit")
            return takeprofit
    except FileNotFoundError:
        print("TakeProfit.json not found. Returning None.")
        return None

def tp_jsonSELL():
    try:
        with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\TakeProfitSELL.json', 'r') as json_file:
            data = json.load(json_file)
            takeProfitSELL = data.get("take_profit")
            return takeProfitSELL
    except FileNotFoundError:
        print("TakeProfitSELL.json not found. Returning None.")
        return None
    
def sp_json():
    try:
        with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\StopLoss.json', 'r') as json_file:
            data = json.load(json_file)
            stoploss = data.get("stop_loss")
            return stoploss
    except FileNotFoundError:
        print("StopLoss.json not found. Returning None.")
        return None

def sp_jsonSELL():
    try:
        with open(r'C:\Users\David\Desktop\Pilot\END_PRODUCT\StopLossSELL.json', 'r') as json_file:
            data = json.load(json_file)
            stopLossSELL = data.get("stop_loss")
            return stopLossSELL
    except FileNotFoundError:
        print("StopLossSELL.json not found. Returning None.")
        return None
'''
# Exemplu:
j_data = read_values_from_j(j_FILE_PATH)
update_excel_and_save_to_j(j_data)
'''
print('STARTING PROFIT PILOT...')
time.sleep(2)

pos = input("Buy position opened(y/n/s): ")
if pos.lower() == 'y':
    pos = 1
if pos.lower() == 'n':
    pos = 0


timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
print(f'start { timestamp }')
lstpos = 0
while True:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    j_data = read_values_from_j(J_FILE_PATH)
    update_excel_and_save_to_j(j_data)

    df = pd.read_excel(EXCEL_FILE_PATH, names=['Date', 'Close'])

    find_optimal_lma_sma(df)

    optimal_lma, optimal_sma, max_sma_lma_price_sum = find_optimal_lma_sma(df)


    #df['SMA'] = df['Close'].rolling(optimal_sma).mean()
    #df['LMA'] = df['Close'].rolling(optimal_lma).mean()

    df['SMA'] = df['Close'].rolling(SMA).mean()
    df['LMA'] = df['Close'].rolling(LMA).mean()
    

    #df['SMAvLMA'] = np.where(df['SMA'] > df['LMA'], 1, 0)
    #df['ClosevMA20'] = np.where(df['Close'] > df['SMA'], 1, 0)
    #df['Conclusion'] = np.where((df['SMAvLMA'] >= 1) & (df['ClosevMA20'] >= 1), 1, 0)
    #df['MACD'] = df['SMA'] - df['LMA']
    #df['SL'] = df['MACD'].rolling(window=9, min_periods=1).mean()
    df['High'] = df['Close'].shift(1).cummax()  
    df['Low'] = df['Close'].shift(1).cummin()
    df['TR'] = np.maximum.reduce([
        df['High'] - df['Low'],
        np.abs(df['High'] - df['Close'].shift(1)),
        np.abs(df['Low'] - df['Close'].shift(1))
    ])
    df['ATR'] = df['TR'].rolling(window=ATRP).mean()

    stop_loss_multiplier = 0.1
    take_profit_multiplier = 0.15   

    df["stopLoss"] = df['LMA'] - stop_loss_multiplier * df['ATR']
    df["takeProfit"] = df['LMA'] + take_profit_multiplier * df['ATR']


    #df["stopLossSELL"] = df["Close"] + SP_P * df["ATR"]
    #df["takeProfitSELL"] = df["Close"] - TP * df["ATR"]

    if pos == 0:
        openPrice = 25000

    #CONDITION:
        
    BuyCondition = False  
    SellCondition = False

    CloseBuy = False
    CloseSell = False


    #CONDITION  - BUY
    #and (df['MACD'].iloc[-1] > df['SL'].iloc[-1]).any()
    if(df['SMA'].iloc[-1] > df['LMA'].iloc[-1]).any()  and pos == 0:
        BuyCondition = True
    

    #CONDITION - CLOSE BUY
    #and (df['MACD'].iloc[-1] < df['SL'].iloc[-1]).any()
    #if (df['SMA'].iloc[-1] < df['LMA'].iloc[-1]).any()  and pos == 1:
    if (df['SMA'].iloc[-1] < df['LMA'].iloc[-1]).any()  and (df['stopLoss'].iloc[-1] >= df['Close'].iloc[-1]) and (df['takeProfit'].iloc[-1] <= df['Close'].iloc[-1]) and pos == 1:
        CloseBuy = True                                                                                                                 


#    BUY
    if BuyCondition is not None and BuyCondition == True:
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]
        stoploss = df['stopLoss'].iloc[-1]
        takeprofit = df['takeProfit'].iloc[-1]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f' BUY - time: {timestamp}, Close: {price}, MAL {optimal_lma} : {lma}, MS {optimal_sma}: {sma} ,SL: {stoploss}, TP: {takeprofit}')
        buy()
        pos = 1
        buyOpenPrice = df['Close'].iloc[-1] 
        BuyCondition = False

#   CONDITION CLOSE BUY   
    if CloseBuy is not None and CloseBuy == True:
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        close()
        ClosePrice = price
        BuyProfit = ((ClosePrice - buyOpenPrice)/ buyOpenPrice)* 100
        print(f'SMA CLOSE ({BuyProfit}) - time: {timestamp}, Close: {price}, MAL {optimal_lma} : {lma}, MS {optimal_sma}: {sma}, SL:{stoploss}') 
        pos = 0
        lstpos = 1
        profit_save(BuyProfit)
        CloseBuy = False

#   OPEN BUY
   # Cprice = df['Close'].iloc[-1]
    if df['SMA'].iloc[-1] > df['LMA'].iloc[-1] and pos == 1:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]  
        pos = 1
        print(f'OPEN BUY - time: {timestamp}, Close: {price}, MAL{optimal_lma} : {lma}, MS {optimal_sma} : {sma}, SL: {stoploss}, TP: {takeprofit} ')

#   NO ACTION
    if pos == 0:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        price = df['Close'].iloc[-1]
        lma = df['LMA'].iloc[-1]
        sma = df['SMA'].iloc[-1]
        print(f'NO ACTION - time: {timestamp}, Close: {price}, MAL {optimal_lma}: {lma}, MS {optimal_sma}: {sma}')
        SELL_reentry_condition = 0
        BUY_reentry_condition = 0


    time.sleep(TIME)
